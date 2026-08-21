"""
Cognos Excel Compiler — produces the professional Excel workbook.

Primary sheet matches the reference workbook format exactly:
    Test Case ID | Category | Test Case Description | Precondition |
    Test Steps | Test Data / Field Reference | Expected Result | Priority

Additional sheets provide full traceability:
    Cover, Requirements, Coverage Summary, Traceability Matrix, Full Details
"""

from __future__ import annotations

import typing
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.drawing.image import Image as OpenpyxlImage  # type: ignore
import os

from app.domain.cognos_test_case import TestSuite
from app.domain.cognos_requirement import RequirementSet, RequirementCategory
from app.domain.reporting_context import FinalReportContext


# Color palette
_COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "alt_row": "F2F7FB",
    "cover_title": "1F4E79",
    "cover_subtitle": "2E75B6",
    "green": "27AE60",
    "amber": "F39C12",
    "red": "E74C3C",
    "light_green": "E8F5E9",
    "light_amber": "FFF8E1",
    "light_red": "FFEBEE",
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _apply_header_style(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color=_COLORS["header_fg"])
        cell.fill = PatternFill(start_color=_COLORS["header_bg"],
                                end_color=_COLORS["header_bg"],
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _THIN_BORDER


def _apply_data_style(ws, row: int, max_col: int, alt: bool = False) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = PatternFill(start_color=_COLORS["alt_row"],
                                    end_color=_COLORS["alt_row"],
                                    fill_type="solid")


def _auto_width(ws, max_col: int, max_width: int = 50) -> None:
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_line = max(len(line) for line in lines) if lines else 0
                    max_len = max(max_len, max_line)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


def build_cognos_workbook(
    context: FinalReportContext,
    generated_at: Optional[datetime] = None,
) -> Workbook:
    """Build the full Cognos UT test case Excel workbook."""
    if generated_at is None:
        generated_at = datetime.now(timezone.utc)

    wb = Workbook()

    # --- Sheet 1: Primary Test Cases ---
    _build_primary_test_sheet(wb, context.test_suite)

    # --- Sheet 2: Cover ---
    _build_cover_sheet(wb, context.test_suite, generated_at)

    # --- Sheet 3: Requirements ---
    _build_requirements_sheet(wb, context)

    # --- Sheet 4: Coverage Summary ---
    _build_coverage_sheet(wb, context.test_suite)

    # --- Sheet 5: Traceability Matrix ---
    _build_traceability_sheet(wb, context)

    # --- Sheet 6: Evidence Snapshots ---
    _build_evidence_snapshots_sheet(wb, context.test_suite)

    return wb


def _build_primary_test_sheet(wb: Workbook, ts: TestSuite) -> None:
    """
    Build the primary Test Scenarios sheet — Phase 10.6 Developer UT format.

    Columns match the reference workbook format:
    Test Case ID | Category | Test Objective | DSD/Technical Reference |
    Preconditions/Test Data | Test Steps | Expected Result | Evidence Required |
    Open Item/Notes | Requirement IDs | Source Page | Source Section |
    LLM Status | Status | Confidence
    """
    ws = typing.cast(typing.Any, wb.active)
    ws.title = "Test Scenarios"
    ws.sheet_properties.tabColor = "2E75B6"

    headers = [
        "Test Case ID",           # A
        "Category",               # B
        "Test Objective",         # C
        "DSD / Technical Reference",  # D
        "Preconditions / Test Data",  # E
        "Test Steps",             # F
        "Expected Result",        # G
        "Evidence Required",      # H
        "Evidence Type",          # I
        "Open Item / Notes",      # J
        "Requirement ID(s)",      # K
        "Source Page",            # L
        "Source Section",         # M
        "LLM Refinement Status",  # N
        "Status",                 # O
        "Confidence",             # P
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ts.test_cases) + 1}"

    # Priority color fills
    _REVIEW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    _OPEN_ITEM_FILL = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")

    for row_idx, tc in enumerate(ts.test_cases, start=2):
        evidences_str = "\n".join(
            [f"- {e.description} ({e.placeholder})" for e in tc.evidence_requirements]
        ) if tc.evidence_requirements else tc.evidence_required

        ev_types = ", ".join(sorted(set(e.evidence_type for e in tc.evidence_requirements))) if tc.evidence_requirements else tc.evidence_type
        req_ids_str = ", ".join(tc.requirement_ids) if tc.requirement_ids else tc.requirement_id

        # Build preconditions + test data as combined cell (reference format)
        prec_data = tc.preconditions
        if tc.test_data and tc.test_data not in ("N/A", "", "N/A — layout verification is structural, no specific data required."):
            prec_data = f"{tc.preconditions}\n\nTest Data:\n{tc.test_data}"

        evidence_pages = ", ".join(sorted(set(str(e.page_number) for e in getattr(tc, 'evidence_references', []) if getattr(e, 'page_number', None))))
        evidence_sections = tc.source_section or "\n".join(sorted(set(e.section for e in getattr(tc, 'evidence_references', []) if e.section)))

        # DSD reference — use new field or fall back to applicability reason
        dsd_ref = getattr(tc, 'dsd_reference', '') or tc.applicability_reason or ""

        # Open item — Phase 10.6 field
        open_item = getattr(tc, 'open_item', '') or tc.open_questions or ""

        # LLM status
        llm_status = getattr(tc, 'llm_refinement_status', 'NOT_ATTEMPTED')

        # Confidence
        confidence = getattr(tc, 'confidence', 'High')

        values = [
            tc.test_case_id,         # A: Test Case ID
            tc.category,             # B: Category
            tc.objective,            # C: Test Objective
            dsd_ref,                 # D: DSD / Technical Reference
            prec_data,               # E: Preconditions / Test Data
            tc.test_steps,           # F: Test Steps
            tc.expected_result,      # G: Expected Result
            evidences_str,           # H: Evidence Required
            ev_types,                # I: Evidence Type
            open_item,               # J: Open Item / Notes
            req_ids_str,             # K: Requirement ID(s)
            evidence_pages,          # L: Source Page
            evidence_sections,       # M: Source Section
            llm_status,              # N: LLM Refinement Status
            tc.status.value if hasattr(tc.status, 'value') else tc.status,  # O: Status
            confidence,              # P: Confidence
        ]

        for col, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=val)

        _apply_data_style(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

        # Highlight rows with open items or REVIEW_REQUIRED status
        status_val = tc.status.value if hasattr(tc.status, 'value') else tc.status
        if open_item:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _OPEN_ITEM_FILL
        elif status_val == "Review Required":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _REVIEW_FILL

    # Set column widths (matching reference workbook proportions)
    col_widths = [16, 28, 45, 40, 45, 55, 50, 40, 18, 35, 22, 14, 28, 20, 18, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights for readability
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(ts.test_cases) + 2):
        ws.row_dimensions[row_idx].height = 60


def _build_cover_sheet(
    wb: Workbook,
    ts: TestSuite,
    generated_at: datetime,
) -> None:
    ws = typing.cast(typing.Any, wb.create_sheet("Cover"))
    ws.sheet_properties.tabColor = _COLORS["cover_title"]

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Cognos Report — Unit Test Case Suite"
    title_cell.font = Font(name="Calibri", size=18, bold=True,
                           color=_COLORS["cover_title"])
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"{ts.report_id} — {ts.report_title}"
    sub.font = Font(name="Calibri", size=14, color=_COLORS["cover_subtitle"])
    sub.alignment = Alignment(horizontal="center")

    meta_rows = [
        ("Report ID", ts.report_id),
        ("Report Name", ts.report_title),
        ("Requirements", str(ts.coverage.total_requirements)),
        ("Test Cases", str(ts.summary.total_generated_ut_cases)),
        ("Covered Requirements", str(ts.coverage.requirements_covered)),
        ("Unmapped", str(ts.coverage.requirements_unmapped)),
        ("Review Required", str(ts.coverage.requirements_ambiguous)),
        ("Coverage %", f"{ts.coverage.overall_coverage_percentage}%"),
    ]

    for idx, (label, value) in enumerate(meta_rows, start=4):
        ws.cell(row=idx, column=1, value=label).font = Font(
            name="Calibri", size=10, bold=bool(label and not value)
        )
        ws.cell(row=idx, column=2, value=value).font = Font(
            name="Calibri", size=10
        )

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _get_traceability_for_req(req, trace_result) -> tuple[str, str, str, str, str]:
    if not trace_result or not req.field:
        return ("", "", "", "", "")

    cat_val = req.category.value if hasattr(req.category, 'value') else str(req.category)
    if cat_val == RequirementCategory.SORT.value:
        for t in trace_result.sort_traces:
            if t.dsd_field_name == req.field:
                mapping = t.mapping_status.value if hasattr(t.mapping_status, 'value') else t.mapping_status
                review = t.review_status.value if hasattr(t.review_status, 'value') else t.review_status
                return (t.xml_field_name, mapping, "", review, "MISSING_IN_XML" if mapping == "MISSING_IN_XML" else "")
    else:
        for t in trace_result.field_traces:
            if t.dsd_field_name == req.field:
                mapping = t.mapping_status.value if hasattr(t.mapping_status, 'value') else t.mapping_status
                impl_type = t.implementation_type.value if hasattr(t.implementation_type, 'value') else t.implementation_type
                review = t.review_status.value if hasattr(t.review_status, 'value') else t.review_status
                notes = "MISSING_IN_XML" if mapping == "MISSING_IN_XML" else ("Complex logic" if getattr(t, "transformation_present", False) else "")
                return (t.xml_data_item_name, mapping, impl_type, review, notes)
    return ("", "", "", "", "")

def _build_requirements_sheet(
    wb: Workbook,
    ctx: FinalReportContext,
) -> None:
    ws = typing.cast(typing.Any, wb.create_sheet("Requirements"))
    ws.sheet_properties.tabColor = "27AE60"

    headers = [
        "Requirement ID",
        "Report ID",
        "Category",
        "Requirement Text",
        "Source Document",
        "Source Section",
        "Source Page",
        "Source Text",
        "Field",
        "Source Table",
        "Source Column(s)",
        "Processing Rule",
        "Formatting Rule",
        "Origin",
        "Status",
        "Mapped Test Case IDs",
        "XML Data Item",
        "XML Mapping Status",
        "Implementation Type",
        "Review Required",
        "Discrepancy Notes",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    req_set = ctx.requirement_set
    ts = ctx.test_suite
    req_items = [r for r in req_set.requirements if not r.is_duplicate_of] if req_set and req_set.requirements else []

    if req_items:
        for row_idx, req in enumerate(req_items, start=2):
            cat_val = req.category.value if hasattr(req.category, 'value') else str(req.category)
            xml_data, xml_mapping, xml_impl, xml_review, xml_notes = _get_traceability_for_req(req, ctx.traceability_result)
            
            # Extract source text from evidence reference if available
            source_text = ""
            if req.evidence_references:
                source_text = req.evidence_references[0].source_text
                
            values = [
                req.requirement_id,
                req.report_id,
                cat_val,
                req.requirement_text,
                req.source_document,
                req.source_section,
                req.source_page or "UNKNOWN",
                source_text,
                req.field or req.business_label,
                req.source_table,
                ", ".join(req.source_columns),
                req.processing_rule,
                req.formatting_rule,
                req.origin,
                req.status,
                ", ".join(req.mapped_test_case_ids),
                xml_data,
                xml_mapping,
                xml_impl,
                xml_review,
                xml_notes,
            ]
            for col, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=val)

            _apply_data_style(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))
    else:
        # Fallback to traceability matrix entries
        for row_idx, entry in enumerate(ts.traceability_matrix, start=2):
            xml_data, xml_mapping, xml_impl, xml_review, xml_notes = "", "", "", "", ""
            if ctx.traceability_result:
                # Find matching req object just to call helper
                req = next((r for r in req_set.requirements if r.requirement_id == entry.requirement_id), None)
                if req:
                    xml_data, xml_mapping, xml_impl, xml_review, xml_notes = _get_traceability_for_req(req, ctx.traceability_result)
                    
            values = [
                entry.requirement_id,
                ts.report_id,
                entry.category,
                entry.requirement_text,
                "",
                "",
                entry.source_page or "UNKNOWN",
                "",  # Source Text
                "",
                "",
                "",
                "",
                "",
                "DIRECT_SPECIFICATION",
                entry.coverage_status,
                ", ".join(entry.test_case_ids),
                xml_data,
                xml_mapping,
                xml_impl,
                xml_review,
                xml_notes,
            ]
            for col, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=val)

            _apply_data_style(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

    _auto_width(ws, len(headers))


def _build_coverage_sheet(wb: Workbook, ts: TestSuite) -> None:
    ws = typing.cast(typing.Any, wb.create_sheet("Coverage Summary"))
    ws.sheet_properties.tabColor = "F39C12"

    headers = [
        "Category",
        "Requirements Found",
        "Requirements Covered",
        "Test Cases Generated",
        "Coverage %",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for row_idx, cat_cov in enumerate(ts.coverage.category_coverage, start=2):
        values = [
            cat_cov.category,
            cat_cov.requirements_found,
            cat_cov.requirements_covered,
            cat_cov.test_cases_generated,
            cat_cov.coverage_percentage,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=val)

        _apply_data_style(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

    total_row = len(ts.coverage.category_coverage) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name="Calibri", size=10, bold=True
    )
    ws.cell(row=total_row, column=2, value=ts.coverage.total_requirements)
    ws.cell(row=total_row, column=3, value=ts.coverage.requirements_covered)
    ws.cell(row=total_row, column=4, value=len(ts.test_cases))
    ws.cell(row=total_row, column=5, value=ts.coverage.overall_coverage_percentage)

    _auto_width(ws, len(headers))


def _build_traceability_sheet(wb: Workbook, ctx: FinalReportContext) -> None:
    ws = typing.cast(typing.Any, wb.create_sheet("Traceability Matrix"))
    ws.sheet_properties.tabColor = "8E44AD"

    headers = [
        "Requirement ID",
        "Requirement Text",
        "Category",
        "Source Page",
        "Test Case ID(s)",
        "Status",
        "XML Data Item",
        "XML Mapping Status",
        "Implementation Type",
        "Review Required",
        "Discrepancy Notes",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    ts = ctx.test_suite
    req_set = ctx.requirement_set
    for row_idx, entry in enumerate(ts.traceability_matrix, start=2):
        xml_data, xml_mapping, xml_impl, xml_review, xml_notes = "", "", "", "", ""
        if ctx.traceability_result:
            req = next((r for r in req_set.requirements if r.requirement_id == entry.requirement_id), None)
            if req:
                xml_data, xml_mapping, xml_impl, xml_review, xml_notes = _get_traceability_for_req(req, ctx.traceability_result)
                
        values = [
            entry.requirement_id,
            entry.requirement_text,
            entry.category,
            entry.source_page or "",
            ", ".join(entry.test_case_ids),
            entry.coverage_status,
            xml_data,
            xml_mapping,
            xml_impl,
            xml_review,
            xml_notes,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=val)

        _apply_data_style(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

    _auto_width(ws, len(headers))

def _build_evidence_snapshots_sheet(wb: Workbook, ts: TestSuite) -> None:
    ws = typing.cast(typing.Any, wb.create_sheet("Evidence Snapshots"))
    ws.sheet_properties.tabColor = "F1C40F"
    
    ws.cell(row=1, column=1, value="Page Number")
    ws.cell(row=1, column=2, value="DSD Snapshot")
    _apply_header_style(ws, 1, 2)
    
    # Collect all unique snapshots
    snapshots = {}
    for tc in ts.test_cases:
        for ref in getattr(tc, "evidence_references", []):
            if ref.snapshot_path and getattr(ref, 'page_number', None) and os.path.exists(ref.snapshot_path):
                snapshots[ref.page_number] = ref.snapshot_path
                
    if not snapshots:
        ws.cell(row=2, column=1, value="No snapshots found.")
        return
        
    current_row = 2
    for page in sorted(snapshots.keys()):
        ws.cell(row=current_row, column=1, value=f"Page {page}")
        img_path = snapshots[page]
        
        try:
            img = OpenpyxlImage(img_path)
            # Resize image if it's too large to prevent Excel crashing or freezing
            max_width = 1000
            if img.width > max_width:
                ratio = max_width / img.width
                img.width = max_width
                img.height = int(img.height * ratio)
                
            ws.add_image(img, f"B{current_row}")
            
            # Estimate height based on image size to adjust row height
            # 1 point is approx 1.33 pixels, adding a little buffer
            ws.row_dimensions[current_row].height = (img.height * 0.75) + 20
        except Exception as e:
            ws.cell(row=current_row, column=2, value=f"Failed to load image: {e}")
            
        current_row += 1
        
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 150
