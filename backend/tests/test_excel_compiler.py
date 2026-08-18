import io

import pytest
from openpyxl import load_workbook

from app.services.excel_compiler import build_workbook

SAMPLE_ROWS = [
    {
        "sl_no": 1, "test_scenario": "Validate SWIPE_CARD_IND values",
        "detailed_test_steps": "1. Query MEMBERS.", "expected_results": "Only Y or N.",
        "verification_sql": "SELECT MEMBERS.SWIPE_CARD_IND FROM MEMBERS;", "source": "ai_generated",
    },
    {
        "sl_no": 2, "test_scenario": "Validate MEMBER_ID uniqueness",
        "detailed_test_steps": "1. Group by MEMBER_ID.", "expected_results": "No duplicates.",
        "verification_sql": "SELECT MEMBER_ID FROM MEMBERS GROUP BY MEMBER_ID HAVING COUNT(*)>1;",
        "source": "ai_generated_edited",
    },
    {
        "sl_no": 3, "test_scenario": "Manual edge case", "detailed_test_steps": "1. Do X.",
        "expected_results": "Y happens.", "verification_sql": "SELECT 1;", "source": "manual",
    },
]

SOURCE_DOCS = [{"filename": "sample_ldm.xlsx", "file_sha256": "a" * 64, "uploaded_at": "2026-07-01T12:00:00"}]


def _roundtrip(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_creates_cover_and_scenarios_sheets():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="desc", source_documents=SOURCE_DOCS)
    reloaded = _roundtrip(wb)
    assert reloaded.sheetnames == ["Cover", "Test Scenarios"]


def test_cover_sheet_has_correct_metadata():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="Some description", source_documents=SOURCE_DOCS)
    reloaded = _roundtrip(wb)
    cover = reloaded["Cover"]
    assert cover["B3"].value == "RPT-Demo"
    assert cover["B4"].value == "CR-1"
    assert cover["B5"].value == "Some description"
    assert cover["B7"].value == "3"  # total scenarios


def test_scenarios_sheet_has_5_required_columns_plus_source():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="d", source_documents=[])
    reloaded = _roundtrip(wb)
    sheet = reloaded["Test Scenarios"]
    header = [sheet.cell(row=1, column=c).value for c in range(1, 7)]
    assert header == ["SL#", "Test Scenario", "Detailed Test Steps", "Expected Results", "Verification SQL", "Source"]


def test_scenario_data_populated_correctly():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="d", source_documents=[])
    reloaded = _roundtrip(wb)
    sheet = reloaded["Test Scenarios"]
    assert sheet.cell(row=2, column=1).value == 1
    assert sheet.cell(row=2, column=2).value == "Validate SWIPE_CARD_IND values"
    assert sheet.cell(row=2, column=5).value == "SELECT MEMBERS.SWIPE_CARD_IND FROM MEMBERS;"


def test_source_labels_are_human_readable():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="d", source_documents=[])
    reloaded = _roundtrip(wb)
    sheet = reloaded["Test Scenarios"]
    assert sheet.cell(row=2, column=6).value == "AI Generated"
    assert sheet.cell(row=3, column=6).value == "AI + Edited"
    assert sheet.cell(row=4, column=6).value == "Manual"


def test_header_row_is_frozen():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="d", source_documents=[])
    reloaded = _roundtrip(wb)
    assert reloaded["Test Scenarios"].freeze_panes == "A2"


def test_empty_rows_raises_value_error():
    with pytest.raises(ValueError):
        build_workbook([], report_id="RPT-X", cr_id=None, cr_description=None)


def test_missing_cr_id_and_description_show_placeholder_text():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id=None, cr_description=None, source_documents=[])
    reloaded = _roundtrip(wb)
    cover = reloaded["Cover"]
    assert cover["B4"].value == "(not provided)"
    assert cover["B5"].value == "(not provided)"


def test_no_source_documents_shows_placeholder_row():
    wb = build_workbook(SAMPLE_ROWS, report_id="RPT-Demo", cr_id="CR-1", cr_description="d", source_documents=[])
    reloaded = _roundtrip(wb)
    cover = reloaded["Cover"]
    # Row 9 is where the doc table body starts given the fixed field layout above
    found_placeholder = any(
        cover.cell(row=r, column=1).value == "(no source documents on record)"
        for r in range(1, cover.max_row + 1)
    )
    assert found_placeholder
