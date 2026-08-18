"""
Excel Compiler — Phase 7.

Turns a session's Interactive Refinement Grid (Phase 6 — the
human-in-the-loop-finalized scenario set, not the raw LLM output) into the
downloadable .xlsx artifact described in the Master System Prompt's Excel
Scenario Output Specification:

  1. SL# — sequential numbering
  2. Test Scenario — high-level functional description
  3. Detailed Test Steps
  4. Expected Results
  5. Verification SQL

A 6th "Source" column is added beyond the spec's minimum 5 — it shows
whether each row is AI Generated, AI + Edited, or Manual (mirroring the
Refinement Grid's own badges), which directly serves the Explainability
requirement (section H: "every generated artifact should include
traceability... mapping back to specific business rules, source
documents"). It's an addition, not a substitution — the 5 required
columns are all present and in the same order the spec lists them.

A separate "Cover" sheet carries Report ID / CR ID / CR Description /
generation timestamp / scenario count, plus a source-document
traceability table (filename + SHA-256 + upload time) — again in service
of Explainability, and reusing the SHA-256 integrity-tracking pattern
already used for uploaded documents (Phase 1) and cached results (Phase 3).

`build_workbook()` is a pure function (rows/metadata in, openpyxl Workbook
out) with no database or filesystem access, specifically so it's testable
without sqlalchemy — see tests/test_excel_compiler.py, which runs for real
in this sandbox (openpyxl is available, sqlalchemy is not).
"""

from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", start_color="1C3BB0", end_color="1C3BB0")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_MONO_FONT = Font(name="Courier New", size=9)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
_THIN_BORDER = Border(*(Side(style="thin", color="D0D5DD") for _ in range(4)))

_SOURCE_LABELS = {
    "ai_generated": "AI Generated",
    "ai_generated_edited": "AI + Edited",
    "manual": "Manual",
}

_TABLE_COLUMNS = [
    # (header, row_key, width, font)
    ("SL#", "sl_no", 6, _BODY_FONT),
    ("Test Scenario", "test_scenario", 32, _BODY_FONT),
    ("Detailed Test Steps", "detailed_test_steps", 45, _BODY_FONT),
    ("Expected Results", "expected_results", 32, _BODY_FONT),
    ("Verification SQL", "verification_sql", 50, _MONO_FONT),
    ("Source", "source", 14, _BODY_FONT),
]


def _style_header_cell(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _style_body_cell(cell, font):
    cell.font = font
    cell.alignment = _WRAP_TOP
    cell.border = _THIN_BORDER


def _build_cover_sheet(wb: Workbook, *, report_id, cr_id, cr_description, generated_at, row_count, source_documents):
    ws = wb.active
    ws.title = "Cover"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    title_font = Font(name="Calibri", bold=True, size=16, color="1C3BB0")
    label_font = Font(name="Calibri", bold=True, size=10)
    value_font = Font(name="Calibri", size=10)

    ws["A1"] = "SIT / QA Test Scenario Export"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    fields = [
        ("Report ID", report_id),
        ("CR ID", cr_id or "(not provided)"),
        ("CR Description", cr_description or "(not provided)"),
        ("Generated At", generated_at.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Total Scenarios", str(row_count)),
    ]
    row = 3
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = value_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Source Documents (traceability)").font = Font(
        name="Calibri", bold=True, size=12, color="1C3BB0"
    )
    row += 1

    doc_headers = ["Filename", "SHA-256", "Uploaded At"]
    for col_idx, header in enumerate(doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _style_header_cell(cell)
    ws.column_dimensions["C"].width = 22
    row += 1

    if not source_documents:
        ws.cell(row=row, column=1, value="(no source documents on record)").font = value_font
    else:
        for doc in source_documents:
            ws.cell(row=row, column=1, value=doc.get("filename", "")).font = value_font
            ws.cell(row=row, column=2, value=doc.get("file_sha256", "")).font = Font(name="Courier New", size=8)
            uploaded_at = doc.get("uploaded_at", "")
            ws.cell(row=row, column=3, value=str(uploaded_at)).font = value_font
            row += 1

    return ws


def _build_scenarios_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Test Scenarios")

    for col_idx, (header, _, width, _) in enumerate(_TABLE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (_, key, _, font) in enumerate(_TABLE_COLUMNS, start=1):
            if key == "source":
                value = _SOURCE_LABELS.get(row_data.get("source"), row_data.get("source", ""))
            else:
                value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_body_cell(cell, font)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_TABLE_COLUMNS))}{len(rows) + 1}"

    return ws


def build_workbook(
    rows: list[dict],
    *,
    report_id: str,
    cr_id: str | None,
    cr_description: str | None,
    source_documents: list[dict] | None = None,
    generated_at: datetime | None = None,
) -> Workbook:
    """
    rows: grid rows as returned by refinement.get_grid() — each needs at
    least sl_no/test_scenario/detailed_test_steps/expected_results/
    verification_sql/source.
    """
    if not rows:
        raise ValueError("Cannot compile an Excel export with zero scenarios.")

    generated_at = generated_at or datetime.now(timezone.utc)

    wb = Workbook()
    _build_cover_sheet(
        wb, report_id=report_id, cr_id=cr_id, cr_description=cr_description,
        generated_at=generated_at, row_count=len(rows),
        source_documents=source_documents or [],
    )
    _build_scenarios_sheet(wb, rows)

    return wb
