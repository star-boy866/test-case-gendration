"""
End-to-end integration test — Phase 10.

Exercises the full service-layer chain in one test, using the real
sample_ldm.xlsx fixture and a fake LLM (no live Ollama needed):

  Ingestion (Phase 1) -> Knowledge Base
    -> Gatekeeper confirmation (Phase 2)
    -> Context Minimizer (Phase 3)
    -> Planning Agent -> AST Builder -> Generator (Phase 4)
    -> Critic -> Reflection Loop (Phase 5)
    -> Refinement Grid (Phase 6)
    -> Excel Compiler (Phase 7)

Deliberately calls SERVICE-LAYER functions directly rather than going
through the FastAPI HTTP layer — `fastapi`/`starlette`'s TestClient isn't
available in the sandbox this was built in (no network to install it), so
this is the deepest integration test achievable without it. It still
proves every phase's actual persistence/business logic works together
against a single shared database and a single shared Knowledge Base,
which is exactly the kind of cross-phase regression an isolated unit test
per module can't catch.

DISCLOSED: this needs sqlalchemy, which also isn't installed in this
sandbox — syntax-checked only here, not executed. Please run this for
real via `pytest tests/test_integration_e2e.py -v` once dependencies are
installed; it is the single highest-value test in this repo to confirm
passes, since it's the only one that exercises every phase together.
"""

from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.session import Base
from app.models import audit, knowledge_base as kb_models, cache, refinement, export, evaluation  # noqa: F401

from app.services.document_parser import parse_document
from app.services.knowledge_base import persist_parsed_document
from app.services.gatekeeper import confirm_scope
from app.services.context_minimizer import minimize_context
from app.agents.pipeline import run_pipeline
from app.agents.reflection_loop import run_reflection_loop
from app.services.refinement import add_generated_rows, get_grid
from app.services.export_service import export_session_to_excel

FIXTURES = Path(__file__).parent / "fixtures"
REPORT_ID = "RPT-Integration"
REQUIREMENT = "Validate Swipe Card Issuance tracking when indicator equals 'N'"


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


def _fake_llm(prompt: str) -> str:
    """
    Deliberately simple, deterministic fake — good enough to drive a real
    Planning->AST->Generator->Critic pass against the sample_ldm.xlsx
    fixture's actual schema (MEMBERS.SWIPE_CARD_IND, MEMBERS.MEMBER_ID,
    CLAIMS.MEMBER_ID/CLAIM_ID), without needing a live LLM.
    """
    import json

    if "JSON array" in prompt:
        # Planning Agent stage: propose scenarios covering BOTH business
        # rules in the fixture (SWIPE_CARD_IND domain + MEMBER_ID PK), so
        # the Critic's business_rules_covered/edge_cases_covered checks
        # can genuinely pass without needing a reflection-loop gap-fill.
        return json.dumps([
            {
                "title": "Validate SWIPE_CARD_IND domain",
                "rationale": "Business rule requires Y or N only",
                "category": "valid_value_check",
                "target_table": "MEMBERS",
                "target_columns": ["SWIPE_CARD_IND"],
                "filters": [{"column": "SWIPE_CARD_IND", "op": "NOT IN", "value": "('Y','N')"}],
            },
            {
                "title": "Validate MEMBER_ID is never null",
                "rationale": "Primary key must never be null",
                "category": "null_check",
                "target_table": "MEMBERS",
                "target_columns": ["MEMBER_ID"],
                "filters": [{"column": "MEMBER_ID", "op": "IS NULL", "value": ""}],
            },
        ])
    # Generator Agent stage: vary output based on which column the prompt
    # is about, like a real LLM would.
    if "SWIPE_CARD_IND" in prompt:
        return json.dumps({
            "test_scenario": "Validate SWIPE_CARD_IND values",
            "detailed_test_steps": "1. Query MEMBERS.SWIPE_CARD_IND for any value outside Y/N.",
            "expected_results": "Zero rows returned.",
        })
    return json.dumps({
        "test_scenario": "Validate MEMBER_ID is never null",
        "detailed_test_steps": "1. Query MEMBERS.MEMBER_ID for nulls.",
        "expected_results": "Zero rows returned.",
    })


def test_full_pipeline_ingestion_through_excel_export(db_session, tmp_path, monkeypatch):
    # --- Phase 1: Ingestion -----------------------------------------------
    parsed = parse_document(FIXTURES / "sample_ldm.xlsx")
    ingest_result = persist_parsed_document(
        db_session,
        report_id=REPORT_ID,
        cr_id="CR-INTEGRATION-1",
        filename="sample_ldm.xlsx",
        file_path=FIXTURES / "sample_ldm.xlsx",
        file_type="xlsx",
        uploaded_by="integration-test",
        parsed=parsed,
    )
    assert ingest_result["parse_status"] == "parsed"
    assert ingest_result["counts"]["columns"] == 4

    # --- Phase 2: Gatekeeper confirmation ----------------------------------
    confirm_result = confirm_scope(
        db_session,
        report_id=REPORT_ID,
        cr_id="CR-INTEGRATION-1",
        cr_description="Integration test CR",
        confirmed_by="integration-test",
    )
    # confirm_scope() signals success by NOT raising GatekeeperError — its
    # return dict has no separate "accepted" boolean to check.
    assert confirm_result["confirmed_by"] == "integration-test"
    session_id = confirm_result["session_id"]

    # --- Phase 3: Context Minimizer ----------------------------------------
    context_slice = minimize_context(db_session, REPORT_ID, REQUIREMENT)
    context_slice_dict = context_slice.to_dict()
    assert "MEMBERS" in context_slice_dict["candidate_tables"]

    # --- Phase 4: Planning Agent -> AST Builder -> Generator ---------------
    generated, pipeline_warnings = run_pipeline(
        context_slice_dict, REQUIREMENT, _fake_llm, max_scenarios=6,
    )
    assert len(generated) == 2, f"expected 2 valid scenarios, got {len(generated)}; warnings={pipeline_warnings}"

    # --- Phase 5: Critic -> Reflection Loop ---------------------------------
    reflection = run_reflection_loop(
        generated, context_slice_dict, REQUIREMENT, _fake_llm, max_iterations=2,
    )
    assert reflection.critic_report.passed is True, reflection.critic_report.issues

    # --- Phase 6: Refinement Grid -------------------------------------------
    scenario_dicts = [s.to_dict() for s in reflection.scenarios]
    add_generated_rows(
        db_session, session_id=session_id, report_id=REPORT_ID,
        requirement_text=REQUIREMENT, scenarios=scenario_dicts,
    )
    grid = get_grid(db_session, session_id)
    assert len(grid) == 2
    assert all(row["source"] == "ai_generated" for row in grid)

    # --- Phase 7: Excel Compiler ---------------------------------------------
    monkeypatch.setattr("app.services.export_service.settings.EXPORT_DIR", str(tmp_path))
    record = export_session_to_excel(db_session, session_id=session_id, exported_by="integration-test")
    assert record.row_count == 2
    assert Path(record.file_path).exists()

    # Confirm the generated file is a genuinely valid, re-openable workbook.
    import openpyxl
    wb = openpyxl.load_workbook(record.file_path)
    assert wb.sheetnames == ["Cover", "Test Scenarios"]
    assert wb["Test Scenarios"].cell(row=1, column=1).value == "SL#"
    assert wb["Test Scenarios"].max_row == 3  # header + 2 scenario rows


def test_full_pipeline_with_partial_business_rule_coverage_triggers_reflection(db_session, tmp_path):
    """
    Same fixture, but the fake LLM's FIRST planning call only covers ONE of
    the two business rules — proving the reflection loop's gap-filling
    (Phase 5) actually engages inside a realistic full-pipeline run, not
    just in reflection_loop.py's own isolated unit tests.
    """
    import json

    parsed = parse_document(FIXTURES / "sample_ldm.xlsx")
    persist_parsed_document(
        db_session, report_id="RPT-Partial", cr_id="CR-2",
        filename="sample_ldm.xlsx", file_path=FIXTURES / "sample_ldm.xlsx",
        file_type="xlsx", uploaded_by="integration-test", parsed=parsed,
    )
    confirm_scope(
        db_session, report_id="RPT-Partial", cr_id="CR-2",
        cr_description="Partial coverage test", confirmed_by="integration-test",
    )
    context_slice_dict = minimize_context(db_session, "RPT-Partial", REQUIREMENT).to_dict()

    def under_covering_llm(prompt):
        if "JSON array" in prompt:
            if "ADDITIONAL INSTRUCTION" in prompt:
                assert "MEMBER_ID" in prompt
                return json.dumps([{
                    "title": "Validate MEMBER_ID is never null", "rationale": "r",
                    "category": "null_check", "target_table": "MEMBERS",
                    "target_columns": ["MEMBER_ID"],
                    "filters": [{"column": "MEMBER_ID", "op": "IS NULL", "value": ""}],
                }])
            return json.dumps([{
                "title": "Validate SWIPE_CARD_IND domain", "rationale": "r",
                "category": "valid_value_check", "target_table": "MEMBERS",
                "target_columns": ["SWIPE_CARD_IND"],
                "filters": [{"column": "SWIPE_CARD_IND", "op": "NOT IN", "value": "('Y','N')"}],
            }])
        if "SWIPE_CARD_IND" in prompt:
            return json.dumps({"test_scenario": "Validate SWIPE_CARD_IND", "detailed_test_steps": "s", "expected_results": "e"})
        return json.dumps({"test_scenario": "Validate MEMBER_ID not null", "detailed_test_steps": "s", "expected_results": "e"})

    generated, _ = run_pipeline(context_slice_dict, REQUIREMENT, under_covering_llm, max_scenarios=6)
    assert len(generated) == 1  # only one rule covered initially

    reflection = run_reflection_loop(generated, context_slice_dict, REQUIREMENT, under_covering_llm, max_iterations=2)
    assert reflection.critic_report.passed is True
    assert len(reflection.scenarios) == 2
    assert reflection.iterations_used == 1
