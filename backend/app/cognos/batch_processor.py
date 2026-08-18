"""
Cognos Report Batch Processor.

Processes multiple Cognos Report Definition documents concurrently with:
- Strict per-report isolation (no cross-contamination of context/state)
- Failure isolation (one failed report does not stop the batch)
- Configurable worker concurrency limits
- Per-report output generation + Batch Summary Excel compilation
"""

from __future__ import annotations

import concurrent.futures
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Sequence
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.cognos.pipeline import run_cognos_pipeline, PipelineResult
from app.services.cognos_excel_compiler import build_cognos_workbook


@dataclass
class ReportResult:
    """Output summary for a single report within a batch."""
    report_id: str
    report_name: str
    source_filename: str
    status: str  # SUCCESS, FAILED, REVIEW_REQUIRED
    requirements_count: int = 0
    test_count: int = 0
    coverage_pct: float = 0.0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    output_excel_path: str = ""
    duration_seconds: float = 0.0


@dataclass
class BatchResult:
    """Overall batch execution summary."""
    batch_id: str
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    reports_processed: int = 0
    reports_successful: int = 0
    reports_failed: int = 0
    report_results: list[ReportResult] = field(default_factory=list)
    batch_summary_excel_path: str = ""


def process_single_report_safe(
    docx_path: str | Path,
    output_base_dir: Path,
) -> ReportResult:
    """
    Process a single report within the batch, catching exceptions to ensure failure isolation.
    """
    start_time = datetime.now()
    path = Path(docx_path)
    source_name = path.name

    try:
        pipeline_res = run_cognos_pipeline(path)
        report_def = pipeline_res.report_definition
        req_set = pipeline_res.requirement_set
        test_suite = pipeline_res.test_suite

        report_id = report_def.metadata.report_id or path.stem
        report_name = report_def.metadata.report_title or "Cognos Report"

        # Build output directory for report
        report_dir = output_base_dir / report_id.replace(" ", "_")
        report_dir.mkdir(parents=True, exist_ok=True)

        # Build and save Excel workbook
        wb = build_cognos_workbook(test_suite, req_set)
        excel_filename = f"{report_id}_UT_Scenarios.xlsx"
        excel_path = report_dir / excel_filename
        wb.save(excel_path)

        duration = (datetime.now() - start_time).total_seconds()
        status = "REVIEW_REQUIRED" if pipeline_res.test_suite.coverage.requirements_ambiguous > 0 else "SUCCESS"

        return ReportResult(
            report_id=report_id,
            report_name=report_name,
            source_filename=source_name,
            status=status,
            requirements_count=len(req_set.requirements),
            test_count=len(test_suite.test_cases),
            coverage_pct=test_suite.coverage.overall_coverage_percentage,
            warnings=test_suite.generation_warnings,
            output_excel_path=str(excel_path),
            duration_seconds=duration,
        )

    except Exception as e:
        duration = (datetime.now() - start_time).total_seconds()
        return ReportResult(
            report_id=path.stem,
            report_name=path.name,
            source_filename=source_name,
            status="FAILED",
            errors=[str(e)],
            duration_seconds=duration,
        )


def process_report_batch(
    docx_paths: Sequence[str | Path],
    output_dir: str | Path | None = None,
    max_workers: int | None = None,
    continue_on_error: bool = True,
) -> BatchResult:
    """
    Process multiple Cognos report documents in parallel with failure isolation.
    """
    batch_id = f"BATCH-{uuid.uuid4().hex[:8].upper()}"

    if output_dir is None:
        output_dir = Path(settings.EXPORT_DIR) / batch_id
    else:
        output_dir = Path(output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    if max_workers is None:
        max_workers = getattr(settings, "MAX_REPORT_WORKERS", 4)

    batch_res = BatchResult(batch_id=batch_id)

    # Concurrency processing
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_path = {
            executor.submit(process_single_report_safe, path, output_dir): path
            for path in docx_paths
        }

        for future in concurrent.futures.as_completed(future_to_path):
            res = future.result()
            batch_res.report_results.append(res)
            batch_res.reports_processed += 1
            if res.status == "FAILED":
                batch_res.reports_failed += 1
                if not continue_on_error:
                    break
            else:
                batch_res.reports_successful += 1

    # Compile Batch Summary Excel
    batch_summary_path = output_dir / f"Batch_Summary_{batch_id}.xlsx"
    _build_batch_summary_excel(batch_res, batch_summary_path)
    batch_res.batch_summary_excel_path = str(batch_summary_path)

    return batch_res


def _build_batch_summary_excel(batch_res: BatchResult, save_path: Path) -> None:
    """Generate a high-level Excel summary workbook for the batch."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Summary"

    headers = [
        "Report ID",
        "Report Name",
        "Source Document",
        "Status",
        "Requirements Found",
        "Test Cases Generated",
        "Coverage %",
        "Duration (s)",
        "Errors / Warnings",
        "Output Path",
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, res in enumerate(batch_res.report_results, start=2):
        err_warn = "; ".join(res.errors + res.warnings[:2])
        values = [
            res.report_id,
            res.report_name,
            res.source_filename,
            res.status,
            res.requirements_count,
            res.test_count,
            res.coverage_pct,
            round(res.duration_seconds, 2),
            err_warn,
            res.output_excel_path,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=val)

    # Widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(save_path)
