"""
Comprehensive Excel Compiler — produces the 7-sheet professional QA workbook.

Sheets:
1. Cover
2. Requirement Summary
3. Coverage Matrix
4. Detailed Test Cases
5. Assumptions & Dependencies
6. Risks & Gaps
7. Methodology
"""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.domain.reporting_context import FinalReportContext
from app.cognos.test_design.comprehensive_test_design_engine import ComprehensiveTestSuite

# Style Setup
_COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "alt_row": "F2F7FB",
    "cover_title": "1F4E79",
    "cover_subtitle": "2E75B6",
}
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def _apply_header_style(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color=_COLORS["header_fg"])
        cell.fill = PatternFill(start_color=_COLORS["header_bg"], end_color=_COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

def _apply_data_style(ws, row: int, max_col: int, alt: bool = False) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = PatternFill(start_color=_COLORS["alt_row"], end_color=_COLORS["alt_row"], fill_type="solid")

def _auto_width(ws, max_col: int, max_width: int = 50) -> None:
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25 # simplified auto width

def build_comprehensive_workbook(
    ctx: FinalReportContext,
    comp_suite: ComprehensiveTestSuite,
    generated_at: Optional[datetime] = None,
) -> Workbook:
    if generated_at is None:
        generated_at = datetime.now(timezone.utc)
        
    wb = Workbook()
    
    # --- 1. Cover ---
    ws_cover = wb.active
    ws_cover.title = "Cover"
    _build_cover_sheet(ws_cover, comp_suite)
    
    # --- 2. Requirement Summary ---
    ws_req = wb.create_sheet("Requirement Summary")
    _build_requirements_sheet(ws_req, ctx)
    
    # --- 3. Coverage Matrix ---
    ws_cov = wb.create_sheet("Coverage Matrix")
    _build_coverage_sheet(ws_cov, comp_suite)
    
    # --- 4. Detailed Test Cases ---
    ws_tc = wb.create_sheet("Detailed Test Cases")
    _build_test_cases_sheet(ws_tc, comp_suite)
    
    # --- 5. Assumptions & Dependencies ---
    ws_assm = wb.create_sheet("Assumptions & Dependencies")
    _build_assumptions_sheet(ws_assm, comp_suite)
    
    # --- 6. Risks & Gaps ---
    ws_risk = wb.create_sheet("Risks & Gaps")
    _build_risks_sheet(ws_risk, comp_suite)
    
    # --- 7. Methodology ---
    ws_meth = wb.create_sheet("Methodology")
    _build_methodology_sheet(ws_meth, comp_suite)
    
    return wb

def _build_cover_sheet(ws, comp_suite: ComprehensiveTestSuite):
    ws.cell(row=1, column=1, value="Comprehensive Test Suite").font = Font(size=18, bold=True)
    ws.cell(row=2, column=1, value=f"Report: {comp_suite.base_suite.report_id}").font = Font(size=14)
    ws.cell(row=4, column=1, value="Total Requirements:")
    ws.cell(row=4, column=2, value=comp_suite.base_suite.coverage.total_requirements)
    ws.cell(row=5, column=1, value="Total Comprehensive Tests:")
    ws.cell(row=5, column=2, value=len(comp_suite.comprehensive_tests) + len(comp_suite.base_suite.test_cases))

def _build_requirements_sheet(ws, ctx: FinalReportContext):
    headers = ["Requirement ID", "Category", "Requirement Text", "Source Document", "Field"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))
    for row, req in enumerate(ctx.requirement_set.requirements, 2):
        ws.cell(row=row, column=1, value=req.requirement_id)
        ws.cell(row=row, column=2, value=req.category.value if hasattr(req.category, 'value') else req.category)
        ws.cell(row=row, column=3, value=req.requirement_text)
        ws.cell(row=row, column=4, value=req.source_document)
        ws.cell(row=row, column=5, value=req.field)
        _apply_data_style(ws, row, len(headers), alt=(row % 2 == 0))
    _auto_width(ws, len(headers))

def _build_coverage_sheet(ws, comp_suite: ComprehensiveTestSuite):
    headers = ["Category", "Requirements Found", "Requirements Covered", "Coverage %"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))
    for row, cov in enumerate(comp_suite.base_suite.coverage.category_coverage, 2):
        ws.cell(row=row, column=1, value=cov.category)
        ws.cell(row=row, column=2, value=cov.requirements_found)
        ws.cell(row=row, column=3, value=cov.requirements_covered)
        ws.cell(row=row, column=4, value=cov.coverage_percentage)
        _apply_data_style(ws, row, len(headers), alt=(row % 2 == 0))
    _auto_width(ws, len(headers))

def _build_test_cases_sheet(ws, comp_suite: ComprehensiveTestSuite):
    headers = [
        "Test Case ID", "Requirement ID", "Test Scenario Title", "Objective",
        "Preconditions", "Test Steps", "Expected Result", "Priority", "Test Type", "Origin"
    ]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))
    
    all_tests = comp_suite.base_suite.test_cases + comp_suite.comprehensive_tests
    
    for row, tc in enumerate(all_tests, 2):
        ws.cell(row=row, column=1, value=tc.test_case_id)
        ws.cell(row=row, column=2, value=tc.requirement_id)
        ws.cell(row=row, column=3, value=tc.test_case_title)
        ws.cell(row=row, column=4, value=tc.objective)
        ws.cell(row=row, column=5, value=tc.preconditions)
        ws.cell(row=row, column=6, value=tc.test_steps)
        ws.cell(row=row, column=7, value=tc.expected_result)
        ws.cell(row=row, column=8, value=tc.priority.value if hasattr(tc.priority, 'value') else tc.priority)
        ws.cell(row=row, column=9, value=tc.test_type.value if hasattr(tc.test_type, 'value') else tc.test_type)
        ws.cell(row=row, column=10, value=tc.origin.value if hasattr(tc.origin, 'value') else tc.origin)
        _apply_data_style(ws, row, len(headers), alt=(row % 2 == 0))
    _auto_width(ws, len(headers))

def _build_assumptions_sheet(ws, comp_suite: ComprehensiveTestSuite):
    headers = ["Assumption ID", "Description", "Dependency", "Status", "Validating Tests"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))
    for row, a in enumerate(comp_suite.assumptions, 2):
        ws.cell(row=row, column=1, value=a.assumption_id)
        ws.cell(row=row, column=2, value=a.description)
        ws.cell(row=row, column=3, value=a.dependency)
        ws.cell(row=row, column=4, value=a.status)
        ws.cell(row=row, column=5, value=", ".join(a.validating_test_case_ids))
        _apply_data_style(ws, row, len(headers), alt=(row % 2 == 0))
    _auto_width(ws, len(headers))

def _build_risks_sheet(ws, comp_suite: ComprehensiveTestSuite):
    headers = ["Risk ID", "Description", "Impact", "Derived From", "Mitigating Tests"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))
    for row, r in enumerate(comp_suite.risks, 2):
        ws.cell(row=row, column=1, value=r.risk_id)
        ws.cell(row=row, column=2, value=r.description)
        ws.cell(row=row, column=3, value=r.impact)
        ws.cell(row=row, column=4, value=r.derived_from)
        ws.cell(row=row, column=5, value=", ".join(r.mitigating_test_case_ids))
        _apply_data_style(ws, row, len(headers), alt=(row % 2 == 0))
    _auto_width(ws, len(headers))

def _build_methodology_sheet(ws, comp_suite: ComprehensiveTestSuite):
    ws.cell(row=1, column=1, value="Test Methodology").font = Font(size=14, bold=True)
    ws.cell(row=3, column=1, value="Description:")
    ws.cell(row=3, column=2, value=comp_suite.methodology.description)
    ws.cell(row=5, column=1, value="Techniques Used:").font = Font(bold=True)
    for r, t in enumerate(comp_suite.methodology.techniques_used, 6):
        ws.cell(row=r, column=1, value=f"- {t}")
    r_idx = 6 + len(comp_suite.methodology.techniques_used) + 1
    ws.cell(row=r_idx, column=1, value="Rules Applied:").font = Font(bold=True)
    for r, rule in enumerate(comp_suite.methodology.rules_applied, r_idx + 1):
        ws.cell(row=r, column=1, value=f"- {rule}")
